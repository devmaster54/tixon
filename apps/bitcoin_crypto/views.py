import json
import requests
import pytz
from decimal import Decimal
from datetime import datetime, timedelta, date
from django.utils import timezone

from django.core import serializers
from django.http import JsonResponse

from django.shortcuts import render, HttpResponse, render_to_response, get_object_or_404, redirect
from django.views.generic import TemplateView, FormView, View, ListView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.utils.translation import ugettext_lazy as _
from django.conf import settings
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_text
from django.core.mail import send_mail, EmailMessage
from django.utils import six
from django.template.loader import render_to_string
from django.urls import reverse
from django.urls import resolve
from django.contrib.sites.models import Site
from django.contrib import messages
from django.db.models import Q, Sum

from apps.bitcoin_crypto.forms import ExchangeForm, ReportForm, TransactionForm, VaultTransactionForm, AddSGDForm, ConfirmFiatTransactionForm, ConfirmCrytpoRequestForm, OrderCoverTransactionForm, DisputeUploadForm
from apps.bitcoin_crypto.models import Transaction, Wallet, PendingTransactions, VaultWallet,\
 VaultTransaction, MinimunCoin, MarketLimit, WatchOnlyAddress, OrderBook, OrderMatchingHistory, SGDWallet, CryptoCurrency, Fiat, Notification, NotificationUser, ConfirmFiatTransaction, ConfirmCrytpoRequest, OrderCoverTransaction, DisputeUpload, PendingOrder
from apps.bitcoin_crypto.utils import *
from apps.authentication.decorators import check_otp, check_2fa
from apps.authentication.models import AccessLog, User, BankAccount, PaypalAccount, WechatAccount, AlipayAccount, Profile
from apps.authentication.forms import BankAccountForm
from django.core.exceptions import PermissionDenied
from django.contrib.auth.mixins import UserPassesTestMixin
from apps.fees.utils import get_transaction_fee
from apps.fees.models import TransactionFee

from apps.authentication.utils import send_otp, _get_pin, send_user_sms, send_admin_sms
from django.contrib.auth.tokens import PasswordResetTokenGenerator
import pyotp

from smtplib import SMTPException

from bitcoinrpc.authproxy import JSONRPCException
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment

from threading import Lock


CURRENCY = {
    '0': 'btc',
    '1': 'eth',
    '2': 'ltc',
    '3': 'xmr',
    '4': 'bch',
    '5': 'btg',

}


class IndexView(TemplateView):
    template_name = 'base.html'


# @method_decorator(login_required, name='dispatch')
# @method_decorator(check_otp, name='dispatch')
# @method_decorator(check_2fa, name='dispatch')
class WelcomeView(TemplateView):
    template_name = 'index-after-login.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        data = self.request.GET
        if not self.request.user.is_anonymous:
            logs = AccessLog.objects.filter(user=self.request.user)
            if len(logs)>2:
                context["last"] = logs[len(logs)-2]

        currencies = CryptoCurrency.objects.all()
        fiats = Fiat.objects.all()
        try:
            cid = int(data['cp'])
        except KeyError:
            cid = currencies[0].id
        try:
            fid = int(data['fp'])
        except KeyError:
            fid = currencies[1].id+100
        base_pair  = currencies.get(pk=cid).basepair
        to_pair  = currencies.get(pk=(fid-100)).basepair if (fid > 100) else currencies.get(pk=fid).basepair
        self.request.session['base_pair'] = base_pair
        self.request.session['to_pair'] = to_pair
        is_otc = 0;
        context["exchange_form"] = ExchangeForm()        
        sell_order = OrderBook.objects.filter(Q(order_type='1'), Q(trade_status=False), Q(amount__gt = F('coins_covered')), Q(exchange_form=base_pair),Q(exchange_to=to_pair)).order_by('price')[:20]
        buy_order = OrderBook.objects.filter(Q(order_type='0'), Q(trade_status=False), Q(amount__gt = F('coins_covered')), Q(exchange_form=base_pair), Q(exchange_to=to_pair)).order_by('-price')[:20]
        if not self.request.user.is_anonymous:
            open_orders = OrderBook.objects.filter(user=self.request.user, trade_status=False).order_by('-order_time')
            order_history = OrderBook.objects.filter(user=self.request.user, trade_status=True).order_by('-order_time')
            context["open_orders"] = open_orders
            context["order_history"] = order_history
        context["sell_order"] = serializers.serialize('json', sell_order)
        context["buy_order"] = serializers.serialize('json', buy_order)
        context["cid"] = cid
        context["fid"] = fid
        context["objectcurrency"] = currencies
        context["fiats"] = fiats
        context["base_pair"] = base_pair
        context["to_pair"] = to_pair
        context["is_otc"] = is_otc
        
        self.request.session['is_otc'] = 0 if fid > 100 else 1

        #context["base_pair"] = currencies[data['pair']-1].split("/")[0]
        #context["to_pair"] = currencies[data['pair']-1].split("/")[1]

        # one year price details
        current_date = datetime.datetime.now()
        date_before_one_year = current_date.replace(year=current_date.year-1)
        price_list = []

        date = date_before_one_year.replace(hour=0,minute=0, second=0, microsecond=0)

        while date.date() <= current_date.date():
            next_date = date + timedelta(days=1)
            try:
                closing_order = OrderMatchingHistory.objects.filter(order_matching_time__range=[date, next_date]).latest('order_matching_time')
            except:
                closing_order = None
            
            if closing_order:
                closing_price = closing_order.matching_price
            elif price_list != []:
                closing_price = price_list[-1][1]
            else:
                date = next_date
                continue

            price_list.append([date.timestamp() * 1000, closing_price])
            date = next_date

        context["price_list"] = price_list

        return context

# @method_decorator(login_required, name='dispatch')
# @method_decorator(check_otp, name='dispatch')
# @method_decorator(check_2fa, name='dispatch')
class OtcWelcomeView(TemplateView):
    template_name = 'index-after-login-otc.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        data = self.request.GET
        if not self.request.user.is_anonymous:
            logs = AccessLog.objects.filter(user=self.request.user)
            if len(logs)>2:
                context["last"] = logs[len(logs)-2]

        currencies = CryptoCurrency.objects.all()
        fiats = Fiat.objects.all()
        try:
            cid = int(data['cp'])
        except KeyError:
            cid = currencies[0].id
        try:
            fid = int(data['fp'])
        except KeyError:
            fid = fiats[0].id
        base_pair  = currencies.get(pk=cid).basepair
        to_pair  = currencies.get(pk=(fid-100)).basepair if (fid > 100) else fiats.get(pk=fid).topair

        self.request.session['base_pair'] = base_pair
        self.request.session['to_pair'] = to_pair
        context["exchange_form"] = ExchangeForm()
        is_otc = 0 if fid > 100 else 1
        if is_otc == 1:
            sell_order = OrderBook.objects.filter(Q(order_type='1'), Q(trade_status=False), Q(amount__gt = F('coins_covered')) , Q(exchange_form=base_pair)|Q(exchange_form=to_pair), Q(exchange_to=base_pair)| Q(exchange_to=to_pair)).select_related('user').order_by('price')[:20]
            buy_order = OrderBook.objects.filter(Q(order_type='0'), Q(trade_status=False), Q(amount__gt = F('coins_covered')), Q(exchange_form=base_pair)|Q(exchange_form=to_pair), Q(exchange_to=base_pair)| Q(exchange_to=to_pair)).select_related('user').order_by('-price')[:20]
        else:
            sell_order = OrderBook.objects.filter(Q(order_type='1'), Q(trade_status=False), Q(amount__gt = F('coins_covered')), Q(exchange_form=base_pair),Q(exchange_to=to_pair)).select_related('user').order_by('price')[:20]
            buy_order = OrderBook.objects.filter(Q(order_type='0'), Q(trade_status=False), Q(amount__gt = F('coins_covered')), Q(exchange_form=base_pair), Q(exchange_to=to_pair)).select_related('user').order_by('-price')[:20]
        if not self.request.user.is_anonymous:
            open_orders = OrderBook.objects.filter(user=self.request.user, trade_status=False).order_by('-order_time')
            order_history = ConfirmFiatTransaction.objects.filter(Q(sender=self.request.user)|Q(receiver=self.request.user), Q(is_confirm=True)).select_related('order')#.order_by('-order_time')
            context["open_orders"] = open_orders
            context["order_history"] = order_history
        # contacts = (so.user.pk for so in sell_order)
        order_users = (o.user_id for o in OrderBook.objects.all())
        pending_order = PendingOrder.objects.filter(Q(is_pending=True)).values('order_id').annotate(sum_pending_amount=Sum('pending_amount'))
        contacts = User.objects.all().filter(Q(id__in = order_users))

        context["sell_order"] = serializers.serialize('json', sell_order)
        context["buy_order"] = serializers.serialize('json', buy_order)
        context["contacts"] = serializers.serialize('json', list(contacts), fields=('phone_number','id'))
        context["cid"] = cid
        context["fid"] = fid
        context["objectcurrency"] = currencies
        context["fiats"] = fiats
        context["base_pair"] = base_pair
        context["to_pair"] = to_pair
        context["is_otc"] = is_otc
        context["pending_orders"] = json.dumps(list(pending_order))
        self.request.session['is_otc'] = 0 if fid > 100 else 1

        #context["base_pair"] = currencies[data['pair']-1].split("/")[0]
        #context["to_pair"] = currencies[data['pair']-1].split("/")[1]

        # one year price details
        current_date = datetime.datetime.now()
        date_before_one_year = current_date.replace(year=current_date.year-1)
        price_list = []

        date = date_before_one_year.replace(hour=0,minute=0, second=0, microsecond=0)

        while date.date() <= current_date.date():
            next_date = date + timedelta(days=1)
            try:
                closing_order = OrderMatchingHistory.objects.filter(order_matching_time__range=[date, next_date]).latest('order_matching_time')
            except:
                closing_order = None
            
            if closing_order:
                closing_price = closing_order.matching_price
            elif price_list != []:
                closing_price = price_list[-1][1]
            else:
                date = next_date
                continue

            price_list.append([date.timestamp() * 1000, closing_price])
            date = next_date

        context["price_list"] = price_list

        return context

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SettingsView(TemplateView):
    template_name = 'settings.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['bank_form'] = BankAccountForm()
        context['bank_accounts'] = BankAccount.objects.filter(user=self.request.user)
        context['paypal_accounts'] = PaypalAccount.objects.filter(user=self.request.user)
        context['wechat_accounts'] = WechatAccount.objects.filter(user=self.request.user)
        context['alipay_accounts'] = AlipayAccount.objects.filter(user=self.request.user)
        return context

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ExchangeRateView(View):
    """
    Displaying the exchange rate based on the amount.
    If no amount entered will take the minimum transaction amount
    """
    def post(self, request, *args, **kwargs):
        convert_from = CURRENCY[request.POST.get('from')]
        convert_to = CURRENCY[request.POST.get('to')]
        amount = request.POST.get('amount')
        
        if amount:
            params =  {
                    "from": convert_from,
                    "to": convert_to,
                    "amount": amount
                }
            method = "getExchangeAmount"
        else:
            params =  {
                    "from": convert_from,
                    "to": convert_to,
                }
            method = 'getMinAmount'
        data = changelly_transaction(method,params)
        if not data.get('error'):
            request.session['convert_from'] = convert_from
            request.session['convert_to'] = convert_to
            request.session['amount'] = amount
            if not amount:
                request.session['amount'] = '1'
        return HttpResponse(json.dumps(data), content_type='application/json')


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class WalletsView(TemplateView):
    template_name = 'bitcoin/wallets.html'

    def get_context_data(self, *args, **kwargs):
        context = super(WalletsView, self).get_context_data(**kwargs)

        #preventing generating new wallet address when page refresh
        # base_pair = self.request.session['base_pair']
        if not Wallet.objects.filter(user=self.request.user, name='BTC').exists():
            create_btc_wallet(self.request.user)
        if not Wallet.objects.filter(user=self.request.user, name='TXCH').exists():
            create_txch_wallet(self.request.user)

        context['wallets'] = Wallet.objects.filter(user=self.request.user)
        return context


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class TransactionListView(ListView):
    """
    All transactions list of a user
    """
    model = Transaction
    template_name = 'bitcoin/transactions.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        access = AuthServiceProxy("http://tixoner:123456789tixon@68.183.176.17:9999")
        transactions = access.listtransactions("*") #self.request.user.username 
        context['btctransactions'] = [txn for txn in transactions if txn['category'] == 'receive']
        context['btcsend_list'] = Transaction.objects.filter(user=self.request.user, currency="BTC", transaction_type='withdrawal')

        access = AuthServiceProxy("http://tixoner:123456789tixon@68.183.185.230:9988")
        transactions = access.listtransactions(self.request.user.username)
        context['txchtransactions'] = [txn for txn in transactions if txn['category'] == 'receive']
        context['txchsend_list'] = Transaction.objects.filter(user=self.request.user, currency="TXCH", transaction_type='withdrawal', pending=False , invalid=False, rejected=False)
        return context


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SendBTransactionView(TemplateView):
    """
        verifing amount and address before sending bitcoin to another address
    """
    template_name = 'bitcoin/send_coin.html'

    def get_context_data(self, *args, **kwargs):
        """
            generating transaction fee list to perform jquery fee amount calculation
        """
        current_url = resolve(self.request.path_info).url_name
        if current_url == "sendbtccoin":
            self.request.session['base_pair'] = 'BTC'
        else :
            self.request.session['base_pair'] = 'TXCH'

        context = super(SendBTransactionView, self).get_context_data(**kwargs)
        try:
            fee_obj = TransactionFee.objects.get(currency=self.request.session['base_pair'], fee_type='withdrawal')
            fee_list = fee_obj.transactionfeerange_set.values_list('value','fees')
        except:
            fee_obj = None
            fee_list = {}

        context['fee_dict'] = json.dumps(dict(fee_list))
        context['fee_obj'] = fee_obj
        return context

    def post(self, request, *args, **kwargs):
        """
            verifing submitter address and amount before confirming withdrawal
        """
        form = TransactionForm(request.POST, user=request.user)
        context = self.get_context_data(**kwargs)

        if not request.user.kyc_verified == 'verified' and not request.user.is_superuser :
            context['error'] = _("Your account has not been activated yet. please submit kyc details and wait for admin approval")

            return render(request, self.template_name, context)

        if form.is_valid():

            request.session['to'] = form.cleaned_data['address']
            request.session['amount'] = request.POST.get('amount')
            context['form'] = form
            context['success'] = True

            return render(request, self.template_name, context)
        else:
            context['form'] = form

            return render(request,self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SendTTransactionView(TemplateView):
    """
        verifing amount and address before sending bitcoin to another address
    """
    template_name = 'bitcoin/send_coin.html'

    def get_context_data(self, *args, **kwargs):
        """
            generating transaction fee list to perform jquery fee amount calculation
        """
        current_url = resolve(self.request.path_info).url_name
        if current_url == "sendbtccoin":
            self.request.session['base_pair'] = 'BTC'
        else :
            self.request.session['base_pair'] = 'TXCH'

        context = super(SendTTransactionView, self).get_context_data(**kwargs)
        try:
            fee_obj = TransactionFee.objects.get(currency=self.request.session['base_pair'], fee_type='withdrawal')
            fee_list = fee_obj.transactionfeerange_set.values_list('value','fees')
        except:
            fee_obj = None
            fee_list = {}

        context['fee_dict'] = json.dumps(dict(fee_list))
        context['fee_obj'] = fee_obj
        return context

    def post(self, request, *args, **kwargs):
        """
            verifing submitter address and amount before confirming withdrawal
        """
        form = TransactionForm(request.POST, user=request.user)

        context = self.get_context_data(**kwargs)

        if not request.user.kyc_verified == 'verified' and not request.user.is_superuser :
            context['error'] = _("Your account has not been activated yet. please submit kyc details and wait for admin approval")

            return render(request, self.template_name, context)

        if form.is_valid():

            request.session['to'] = form.cleaned_data['address']
            request.session['amount'] = request.POST.get('amount')
            context['form'] = form
            context['success'] = True

            return render(request, self.template_name, context)
        else:
            context['form'] = form

            return render(request,self.template_name, context)


class EmailTokenGenerator(PasswordResetTokenGenerator):
    """ Overriding default Password reset token generator for email confirmation"""
    def _make_hash_value(self, user, timestamp):
        return (six.text_type(user.pk) + six.text_type(timestamp)) +  six.text_type(user.is_active)

email_token = EmailTokenGenerator()


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SendConfirmView(View):
    def post(self, request, *args, **kwargs):
        
        sms_redirect = False

        if request.POST.get('to') and request.POST.get('amount'):
            request.session['address'] = request.POST.get('to')
            request.session['amount'] = request.POST.get('amount')
            currency = request.POST.get('cointype')
            try:
                del request.session['withdraw-otp-verified']
            except:
                pass

        if request.user.sms_otp:

            otp_verified = request.session.get('withdraw-otp-verified', False)
            
            if not otp_verified and  self.request.POST.get('value') and not request.POST.get('resend-otp'):
                if self.request.POST.get('value') == self.request.session['withdraw-otp']:
                    request.session['withdraw-otp-verified'] = True
                    sms_redirect = True
                else:
                    return HttpResponse(json.dumps({"status":"otp-not-verified"}), content_type='application/json')

            elif not otp_verified:
                pin =  _get_pin(self)
                self.request.session['withdraw-otp'] = pin
                send_otp(self, pin, self.request.user.phone_number)
                return HttpResponse(json.dumps({"status":"get-otp"}), content_type='application/json')

        if request.user.google_2fa:

            if self.request.POST.get('value') and not sms_redirect:
                totp_code = self.request.POST.get('value', None)
                totp = pyotp.TOTP(self.request.user.google_2fa_key)

                if not totp.verify(totp_code):
                    return HttpResponse(json.dumps({"status":"totp-not-verified"}), content_type='application/json')
            else:
                return HttpResponse(json.dumps({"status":"get-totp"}), content_type='application/json')

        transaction_fee = get_transaction_fee(request.session['amount'], currency, 'withdrawal')
        pending_transaction = PendingTransactions.objects.create(user=request.user, amount=request.session['amount'],
            currency=currency, transaction_to=request.session['address'], transaction_fee=str(transaction_fee))

        token = email_token.make_token(request.user)
        uidb64 = urlsafe_base64_encode(force_bytes(pending_transaction.pk)).decode("utf-8")

        html_message = render_to_string('bitcoin/confirm-withdrawal-email.html', {
            'pending_transaction': pending_transaction,
            'uri': reverse('coins:confirm_withdrawal', kwargs={'uidb64':uidb64, 'token': token}),
            'domain': self.request.scheme+"://"+"tixon.exchange",
        })

        try:
            del request.session['withdraw-otp-verified']
        except:
            pass

        try:
            # email = EmailMessage('Subject', 'Body', to=[request.user.email])
            # email.send();
            send_mail('Confirm Withdrawal',
                '',
                settings.DEFAULT_FROM_EMAIL,
                [request.user.email],
                html_message = html_message,
                fail_silently=False
            )
            return HttpResponse(json.dumps({"success":True}), content_type='application/json')
        except:
            return HttpResponse(json.dumps({"error":_("something went wrong")+request.user.email+"  "+settings.DEFAULT_FROM_EMAIL}), content_type='application/json')


class SendEmailConfirmView(View):

    def get(self, request, *args, **kwargs):
        pk = force_text(urlsafe_base64_decode( kwargs.get('uidb64')))
        token = kwargs.get('token')
        try:
            pending_transaction = get_object_or_404(PendingTransactions, pk=pk)
            
        except:
            context={'error' : _('Transaction already completed or not a valid link')}

            return render(request,'bitcoin/transaction_confirm.html', context)

        context = {'success': False}

        if email_token.check_token(pending_transaction.user, token):

            balance = get_balance(pending_transaction.user, pending_transaction.currency)
            access = AuthServiceProxy("http://tixoner:123456789tixon@68.183.176.17:9999") if pending_transaction.currency == 'BTC'else AuthServiceProxy("http://tixoner:123456789tixon@68.183.185.230:9988")
            address = pending_transaction.transaction_to
            amount = Decimal(pending_transaction.amount)
            transaction_fee = Decimal(pending_transaction.transaction_fee)

            if balance < (amount+transaction_fee):
                context = {'message': 'Insufficient Balance.'}
                return render(request,'bitcoin/transaction_confirm.html', context)

            btc_limit_obj, create = MinimunCoin.objects.get_or_create(site=Site.objects.get_current())
            minimum_limit = Decimal(btc_limit_obj.btc_limit)

            balance_status = get_account_balance() - (amount+transaction_fee)
            if (get_account_balance() < (amount+transaction_fee)) or (balance_status < minimum_limit):
                balance = balance - (amount+transaction_fee)
                
                if (get_account_balance() < (amount+transaction_fee)):
                    message = "User requesting more amount than existing balance. Please Visit TixonExchange Wallet Update page to view the Amount need to maintain mininum Balance."
                    email_title = 'Insufficient Fund in Hot Wallet'
                    context = {'success': True}
                else:
                    message = 'Minimum Balance limit Exceed in TixonExchange. Please Update wallet Amount'
                    email_title = 'Minimum Balance limit Exceed'

                valid = access.validateaddress(address)

                if valid['isvalid']:
                    Transaction.objects.create(user=pending_transaction.user, currency=pending_transaction.currency, balance=balance, transaction_type='withdrawal',
                   amount=amount, transaction_fee=str(transaction_fee), transaction_id='', transaction_to=address, pending=True)
                    context = {'success': True}
                else:
                    context = {'message': 'Transaction Failed. Invalid Address'}
                            
                pending_transaction.delete()

                if btc_limit_obj.low_limit_alert:

                    users = User.objects.filter(is_superuser=True).exclude(phone_number="", email="")

                    for admin_user in users:

                        send_mail(email_title,
                            message,
                            settings.DEFAULT_FROM_EMAIL,
                            [admin_user.email],
                            fail_silently=True
                        )

                    btc_limit_obj.low_limit_alert = False
                    btc_limit_obj.save()

                return render(request,'bitcoin/transaction_confirm.html', context)

            valid = access.validateaddress(address)

            if valid['isvalid']:
                balance = balance - (amount+transaction_fee)

                Transaction.objects.create(user=pending_transaction.user, currency=pending_transaction.currency, balance=balance, transaction_type='withdrawal', 
                   amount=amount, transaction_fee=str(transaction_fee), transaction_id='', transaction_to=address, pending=True)
                pending_transaction.delete()
                context = {'success': True}

        return render(request,'bitcoin/transaction_confirm.html', context)


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class PendingTransactionListView(UserPassesTestMixin, View):

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        pending_transactions = Transaction.objects.filter(pending=True)
        context = {
            'pending_transactions': pending_transactions
        }
        return render(request,'bitcoin/pending_transactions.html', context)


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ConfirmPendingTransactionView(UserPassesTestMixin, View):

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):

        try:
            transaction = get_object_or_404(Transaction, pk=self.kwargs['pk'], pending=True)
        except:
            error = _('Transaction object not found')

        try:
            status = complete_pending_transaction(transaction)
            
            if status == True:
                notification = Notification.objects.create(
                    notification='Withdrawal amount %s to address %s has approved by admin'
                    %(transaction.amount, transaction.transaction_to))
                NotificationUser.objects.create(notification=notification, user=transaction.user)
            else :
                error = status

        except JSONRPCException as e:
            status = False
            error = e.error['message']


        if status == True:
            return HttpResponse(json.dumps({"success": True}), content_type='application/json')

        return HttpResponse(json.dumps({"error": error}), content_type='application/json')


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class RejectPendingTransactionView(UserPassesTestMixin, View):
    template_name = 'bitcoin/reject_trnasaction.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        transaction = get_object_or_404(Transaction, pk=self.kwargs['pk'], pending=True)
        context ={
            'transaction': transaction
        }           
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):

        feedback = request.POST.get('feedback', False)
        transaction = get_object_or_404(Transaction, pk=self.kwargs['pk'], pending=True)

        if not feedback:
            context ={
                'transaction': transaction,
                'feedback_error': _('This field is required')
            }  
            return render(request, self.template_name, context)

        transaction.rejected = True
        transaction.pending = False
        transaction.save()

        notification = Notification.objects.create(
            notification='Withdrawal amount %s to address %s has rejected by admin<br>%s'
            %(transaction.amount, transaction.transaction_to, feedback))
        NotificationUser.objects.create(notification=notification, user=transaction.user)

        context ={
            'transaction': transaction,
            'success': True
        }         
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class GetAllUserAddress(TemplateView):
    template_name = 'bitcoin/user_accounts.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        # wallet_object = Wallet.objects.all()
        # context['wallet_obj'] = wallet_object
        
        return render(request,self.template_name, context)


class TransactionRecordView(TemplateView):

    def get(self, request, *args, **kwargs):
        access = create_connection()
        withdraw_transactions = Transaction.objects.all().exclude(transaction_type='fee_withdrawal')

        filtered_deposit = []
        for user in User.objects.all():
            deposit_transaction = access.listtransactions("*") #user.username
            for trnas in deposit_transaction:
                if trnas['category'] == 'receive':
                    filtered_deposit.append(trnas)


        return self.render_to_response({
            'withdraw_transactions':withdraw_transactions,
            'deposit_transaction':filtered_deposit
        })


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SetMinimunLimitBTC(UserPassesTestMixin, TemplateView):
    template_name = 'bitcoin/set_minimum_limit.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def post(self, request, *args, **kwargs):
        value = request.POST.get('minimum_limit')

        try:
            decimal_value = Decimal(value)
        except:
            return HttpResponse(json.dumps({"error": _("Invalid Amount")}), content_type='application/json')

        min_value_obj, create = MinimunCoin.objects.get_or_create(site=Site.objects.get_current())
        min_value_obj.btc_limit = value
        min_value_obj.save()
        return HttpResponse(json.dumps({"success": True}), content_type='application/json')

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SetMarketLimit(UserPassesTestMixin, TemplateView):
    template_name = 'bitcoin/set_market_limit.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):

        market_value_obj, create = MarketLimit.objects.get_or_create()

        context ={
            'market_value': market_value_obj
        }

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        min_price_limit = request.POST.get('min_price_limit')
        max_price_limit = request.POST.get('max_price_limit')
        min_amount_limit = request.POST.get('min_amount_limit')
        max_amount_limit = request.POST.get('max_amount_limit')

        try:
            decimal_min_price_limit = Decimal(min_price_limit)
            decimal_max_price_limit = Decimal(max_price_limit)
            decimal_min_amount_limit = Decimal(min_amount_limit)
            decimal_max_amount_limit = Decimal(max_amount_limit)
            if decimal_min_price_limit > decimal_max_price_limit:
                return HttpResponse(json.dumps({"error": _("Invalid Price Setting")}), content_type='application/json')
            if decimal_min_amount_limit > decimal_max_amount_limit:
                return HttpResponse(json.dumps({"error": _("Invalid Amount Setting")}), content_type='application/json')
        except:
            return HttpResponse(json.dumps({"error": _("Invalid Amount")}), content_type='application/json')

        market_value_obj, create = MarketLimit.objects.get_or_create()
        market_value_obj.min_price_limit = min_price_limit
        market_value_obj.max_price_limit = max_price_limit
        market_value_obj.min_amount_limit = min_amount_limit
        market_value_obj.max_amount_limit = max_amount_limit
        market_value_obj.save()
        return HttpResponse(json.dumps({"success": True}), content_type='application/json')

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SolveDispute(View):
    template_name = 'bitcoin/dispute.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):

        market_value_obj, create = MarketLimit.objects.get_or_create()

        context ={
            'market_value': market_value_obj
        }

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        from_address = request.POST.get('from_address')
        to_address = request.POST.get('to_address')
        amount = request.POST.get('amount')
        valid = transfer_coin(from_address, to_address, amount, 'TXCH')
        if valid:
            return HttpResponse(json.dumps({"success": True}), content_type='application/json')
        return HttpResponse(json.dumps({"error": "Invalid Address"}), content_type='application/json')            
        # try:
        #     decimal_min_price_limit = Decimal(min_price_limit)
        #     decimal_max_price_limit = Decimal(max_price_limit)
        #     decimal_min_amount_limit = Decimal(min_amount_limit)
        #     decimal_max_amount_limit = Decimal(max_amount_limit)
        #     if decimal_min_price_limit > decimal_max_price_limit:
        #         return HttpResponse(json.dumps({"error": _("Invalid Price Setting")}), content_type='application/json')
        #     if decimal_min_amount_limit > decimal_max_amount_limit:
        #         return HttpResponse(json.dumps({"error": _("Invalid Amount Setting")}), content_type='application/json')
        # except:
        #     return HttpResponse(json.dumps({"error": _("Invalid Amount")}), content_type='application/json')

        # market_value_obj, create = MarketLimit.objects.get_or_create()
        # market_value_obj.min_price_limit = min_price_limit
        # market_value_obj.max_price_limit = max_price_limit
        # market_value_obj.min_amount_limit = min_amount_limit
        # market_value_obj.max_amount_limit = max_amount_limit
        # market_value_obj.save()

        # return HttpResponse(json.dumps({"success": True}), content_type='application/json')

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class RaiseDisputeView(View):

    template_name = "bitcoin/raise_dispute.html"

    def get(self, request, *args, **kwargs):
        """
            provide kyc forms to provide kyc details
        """
        dispute_form = DisputeUploadForm()

        context = {
            'dispute_form': dispute_form,
            'timezones': pytz.common_timezones
        }

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """
            kyc details uploading verifing the data and reconfirm by user
        """

        dispute_form = DisputeUploadForm(request.POST, request.FILES)

        confirmation = request.POST.get('confirmation')
        timezone = request.POST.get('timezone')

        if dispute_form.is_valid():

            dispute = dispute_form.save(commit=False)
            dispute.user = self.request.user
            dispute.dispute_status = "processing"
            dispute.save()

            context = {
                'dispute_form': DisputeUploadForm(instance=dispute),
                'confirm_data' : True,
                'timezones': pytz.common_timezones
            }
            return render(request, self.template_name, context)
        else:
            context = {
                'dispute_form': dispute_form,
                'timezones': pytz.common_timezones
            }
            return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ConfirmDisputeView(View):

    def get(self, request, *args, **kwargs):

        notification = Notification.objects.create(notification="New Dispute submission by %s" %request.user.username)
        for user in User.objects.filter(is_superuser=True):
            NotificationUser.objects.create(user=user, notification=notification)

        return redirect(reverse('coins:raise_dispute'))

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ListDisputeView(UserPassesTestMixin, View):

    template_name = "bitcoin/list_dispute.html"

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        """
            provide dispute forms to provide dispute details
        """
        dispute_list = DisputeUpload.objects.filter(dispute_status='processing')

        context = {
            'dispute_list': dispute_list,
            'processing': True
        }

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ListDisputeProcessingView(UserPassesTestMixin, View):

    template_name = "bitcoin/list_dispute.html"

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        
        dispute_list = DisputeUpload.objects.filter(dispute_status='processing')

        context = {
            'dispute_processing_list': dispute_list,
            'processing': True
        }

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ListDisputeApproveView(UserPassesTestMixin, View):

    template_name = "bitcoin/list_dispute.html"

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        """
            provide dispute forms to provide dispute details
        """
        dispute_list = DisputeUpload.objects.filter(dispute_status='approved')

        context = {
            'dispute_approved_list': dispute_list,
            'approved': True
        }

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ListDisputeRejectedView(UserPassesTestMixin, View):

    template_name = "bitcoin/list_dispute.html"

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        
        dispute_list = DisputeUpload.objects.filter(dispute_status='rejected')

        context = {
            'dispute_rejected_list': dispute_list,
            'rejected': True
        }

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class DisputeDetailView(View):

    template_name = "bitcoin/dispute_detail.html"

    def get(self, request, *args, **kwargs):
        disputeDetail = DisputeUpload.objects.get(pk=self.kwargs['pk'])
        profile = Profile.objects.get(user_id=disputeDetail.user_id)
        user = User.objects.get(pk=disputeDetail.user_id)
        try:
            client = User.objects.get(phone_number=disputeDetail.client_phonenumber)
            client_profile = Profile.objects.get(user_id=client.pk)
            context = {
                'disputeDetail': disputeDetail,
                'user_profile': profile,
                'user': user,
                'client': client,
                'client_profile': client_profile,
                'client_invalid': False
            }
        except:
            context = {
                'disputeDetail': disputeDetail,
                'user_profile': profile,
                'user': user,
                'client_invalid': True
            }
        # client = get_object_or_404(User, phone_number=disputeDetail.client_phonenumber)

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class DisputeApproveView(UserPassesTestMixin, View):

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        
        disputeUpload = DisputeUpload.objects.get(pk=self.kwargs['pk'])
        disputeUpload.dispute_status = 'approved'
        disputeUpload.save()

        notification = Notification.objects.create(notification=_("Dispute details approved by admin."))
        NotificationUser.objects.create(user=disputeUpload.user, notification=notification)

        messages.success(self.request, '%s Dispute details Verified.' %disputeUpload.user_email)

        return redirect(reverse('coins:list_dispute'))

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class DisputeRejectView(UserPassesTestMixin, View):

    template_name = 'bitcoin/reject_dispute.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        disputeUpload = get_object_or_404(DisputeUpload, pk=self.kwargs['pk'])
        profile = Profile.objects.get(user_id=disputeUpload.user_id)
        user = User.objects.get(pk=disputeUpload.user_id)
        context = {
            'disputeUpload' : disputeUpload,
            'user_profile': profile,
            'user': user
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):

        feedback = request.POST.get('feedback', False)
        disputeUpload = get_object_or_404(DisputeUpload, pk=self.kwargs['pk'])

        if not feedback:
            context ={
                'disputeUpload': disputeUpload,
                'feedback_error': 'This field is required'
            }
            return render(request, self.template_name, context)

        disputeUpload.dispute_status = 'rejected'
        disputeUpload.save()

        notification = Notification.objects.create(notification=
                                                   _("<b>Dispute details rejected by admin</b><br>%s<br>please contact site administrator for more details") %feedback)
        NotificationUser.objects.create(user=disputeUpload.user, notification=notification)

        messages.success(self.request, _('%s Dispute details Rejected.') %disputeUpload.user_email)
        return redirect(reverse('coins:list_dispute'))

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class DisputeResubmitView(View):

    template_name = "bitcoin/dispute_resubmission.html"

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        disputeUpload = get_object_or_404(DisputeUpload, pk=self.kwargs['pk'])
        profile = Profile.objects.get(user_id=disputeUpload.user_id)
        user = User.objects.get(pk=disputeUpload.user_id)
        context = {
            'disputeUpload' : disputeUpload,
            'user_profile': profile,
            'user': user
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):

        feedback = request.POST.get('feedback', False)
        disputeUpload = get_object_or_404(DisputeUpload, pk=self.kwargs['pk'])

        if not feedback:
            context ={
                'disputeUpload': disputeUpload,
                'feedback_error': 'This field is required'
            }
            return render(request, self.template_name, context)

        disputeUpload.dispute_status = 'rejected'
        disputeUpload.save()

        notification = Notification.objects.create(notification=
                                                   _("<b>Dispute details resubmission require accessing website features</b><br>%s<br>More information contact administrator") %feedback)
        NotificationUser.objects.create(user=disputeUpload.user, notification=notification)

        messages.success(self.request, _('%s Dispute details resubmission requested.') %disputeUpload.user_email)
        return redirect(reverse('coins:list_dispute'))

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class MyListDisputeView(View):

    template_name = "bitcoin/my_list_dispute.html"

    def get(self, request, *args, **kwargs):
        """
            provide dispute forms to provide dispute details
        """
        dispute_list = DisputeUpload.objects.filter(user=self.request.user)

        context = {
            'dispute_list': dispute_list,
        }

        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class WalletToVault(UserPassesTestMixin, TemplateView):
    template_name = 'bitcoin/send_to_vault.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def post(self, request, *args, **kwargs):

        context = self.get_context_data(**kwargs)

        if not request.user.email:
            context['error'] = "For Vault transaction, User need to Add a Valid Email Address."
            return render(request, self.template_name, context)

        # if not request.user.sms_otp and not request.user.google_2fa:
        #     context['error'] = "For Vault transaction, User need to Enable google authentication and sms otp verification."
        #     return render(request, self.template_name, context)

        form = VaultTransactionForm(request.POST, user=request.user)

        if form.is_valid():

            request.session['to'] = form.cleaned_data['address']
            request.session['amount'] = request.POST.get('amount')

            # try:
            #     del request.session['withdraw-otp-verified']
            # except:
            #     pass

            # pin =  _get_pin(self)
            # self.request.session['withdraw-otp'] = pin
            # send_otp(self, pin, self.request.user.phone_number)

            # return redirect(reverse('coins:confirm_wallet_to_vault'))

            """
                code only for testing starts
            """
            amount = request.session['amount']
            address = request.session['to']

            pending_transaction = PendingTransactions.objects.create(user=request.user, amount=amount,
            currency=request.session['base_pair'], transaction_to=address)
            token = email_token.make_token(request.user)
            uidb64 = urlsafe_base64_encode(force_bytes(pending_transaction.pk)).decode("utf-8")

            html_message = render_to_string('bitcoin/confirm-withdrawal-email.html', {
                'pending_transaction': pending_transaction,
                'domain': self.request.scheme+"://"+"tixon.exchange",
                'uri': reverse('coins:vault_confirm', kwargs={'uidb64':uidb64, 'token': token}),
            })

            try:
                send_mail('Confirm Withdrawal',
                    '',
                    settings.DEFAULT_FROM_EMAIL,
                    [request.user.email],
                    html_message = html_message,
                    fail_silently=False
                )
            except:
                pass
                
            context['success'] = True
            """
                code only for testing ends
            """

        context['form'] = form
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ConfirmWalletToVault(View):
    template_name = 'bitcoin/send_to_vault.html'

    def get(self, request, *args, **kwargs):

        form = VaultTransactionForm(initial={'address': request.session['to'],
            'amount': request.session['amount']}, user=request.user)

        context ={
        'sms_verification': True,
        'form': form
        }

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):

        sms_redirect = False
        otp_verified = request.session.get('withdraw-otp-verified', False)
        
        if not otp_verified and  self.request.POST.get('value') and not request.POST.get('resend-otp'):
            if self.request.POST.get('value') == self.request.session['withdraw-otp']:
                request.session['withdraw-otp-verified'] = True
                sms_redirect = True
            else:
                return HttpResponse(json.dumps({"status":"otp-not-verified"}), content_type='application/json')

        elif not otp_verified:
            pin =  _get_pin(self)
            self.request.session['withdraw-otp'] = pin
            send_otp(self, pin, self.request.user.phone_number)
            return HttpResponse(json.dumps({"status":"get-otp"}), content_type='application/json')


        if self.request.POST.get('value') and not sms_redirect:
            totp_code = self.request.POST.get('value', None)
            totp = pyotp.TOTP(self.request.user.google_2fa_key)

            if not totp.verify(totp_code):
                return HttpResponse(json.dumps({"status":"totp-not-verified"}), content_type='application/json')
        else:
            return HttpResponse(json.dumps({"status":"get-totp"}), content_type='application/json')


        amount = request.session['amount']
        address = request.session['to']

        pending_transaction = PendingTransactions.objects.create(user=request.user, amount=amount,
        currency=request.session['base_pair'], transaction_to=address)
        token = email_token.make_token(request.user)
        uidb64 = urlsafe_base64_encode(force_bytes(pending_transaction.pk)).decode("utf-8")

        html_message = render_to_string('bitcoin/confirm-withdrawal-email.html', {
            'pending_transaction': pending_transaction,
            'domain': self.request.scheme+"://"+"tixon.exchange",
            'uri': reverse('coins:vault_confirm', kwargs={'uidb64':uidb64, 'token': token}),
        })

        try:
            send_mail('Confirm Withdrawal',
                '',
                settings.DEFAULT_FROM_EMAIL,
                [request.user.email],
                html_message = html_message,
                fail_silently=False
            )

            return HttpResponse(json.dumps({"success": True}), content_type='application/json')
        except:
            error = _("Email can not send please check your email address and try again")
            return HttpResponse(json.dumps({"error": error}), content_type='application/json')


class WalletToVaultEmailConfirm(View):

    def get(self, request, *args, **kwargs):
        pk = force_text(urlsafe_base64_decode( kwargs.get('uidb64')))
        token = kwargs.get('token')
        try:
            pending_transaction = get_object_or_404(PendingTransactions, pk=pk)
        except:
            context = {'error' : 'Vault transaction already completed or not a valid link' }
            return render(request,'bitcoin/vault_transaction_confirm.html', context)
            
        context = {'success': False}

        if email_token.check_token(pending_transaction.user, token):

            balance = get_account_balance()
            access = create_connection()
            address = pending_transaction.transaction_to
            amount = Decimal(pending_transaction.amount)

            btc_limit_obj, create = MinimunCoin.objects.get_or_create(site=Site.objects.get_current())
            minimum_limit = Decimal(btc_limit_obj.btc_limit)
            if (balance - amount) < minimum_limit:
                context = {
                    'message' : 'Cancelled!. Transaction will not maintain Minimun Balance Limit.'
                }
                return render(request,'bitcoin/vault_transaction_confirm.html', context)

            if balance >= amount:

                try:
                    access.importaddress(address,'',False)
                except:
                    context = {
                        'message' : 'Canceled!. The wallet already contains the private key for this address'
                    }
                    return render(request,'bitcoin/vault_transaction_confirm.html', context)

                valid = access.sendtoaddress(address, amount)

                if valid:
                    WatchOnlyAddress.objects.get_or_create(address=address)
                    txn_fee = access.gettransaction(valid)['fee']
                    balance = balance - amount
                    VaultTransaction.objects.create(user=pending_transaction.user, currency="btc", balance=balance, 
                        amount=amount, mining_fee=str(abs(txn_fee)), transaction_id=valid, transaction_to=address)
                    pending_transaction.delete()
                    context = {'success': True}

        return render(request,'bitcoin/vault_transaction_confirm.html', context)


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class VaultTransactionListView(UserPassesTestMixin, View):

    template_name = 'bitcoin/vault_transactions.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):
        access = create_connection()
        withdraw_transactions = VaultTransaction.objects.filter(transaction_type='to_vault').order_by('-date')
        deposit_transaction =  VaultTransaction.objects.filter(transaction_type='from_vault').order_by('-date')
        context = {
            'withdraw_transactions':withdraw_transactions,
            'deposit_transaction':deposit_transaction
        }
        return render(request, self.template_name, context)




@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ListVaultView(UserPassesTestMixin, View):

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def get(self, request, *args, **kwargs):

        checkwatchonlyaddress()
        vault_list = WatchOnlyAddress.objects.all()
        
        context = {
            'vault_list': vault_list
        }
        return render(request, 'bitcoin/vault_list.html', context)


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class VaultToWalletView(UserPassesTestMixin, TemplateView):
    """send coin back to requested users"""
    template_name = 'bitcoin/vault_to_wallet.html'

    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def post(self, request, *args, **kwargs):
        vault_address = request.POST.get('vault_address')

        if not request.user.email:
            return HttpResponse(json.dumps({"error": "For Vault transaction, User need to Add a Valid Email Address."}), content_type='application/json')

        # if not request.user.sms_otp and not request.user.google_2fa:
        #     return HttpResponse(json.dumps({"error": "For Vault transaction, User need to Enable google authentication and sms otp verification."}), content_type='application/json')


        sms_redirect = False
        if vault_address:

            request.session['address'] = vault_address

            try:
                del request.session['vault2wallet-otp-verified']
            except:
                pass

        # if request.user.sms_otp:
        #     print('sms otp')
        #     otp_verified = request.session.get('vault2wallet-otp-verified', False)
            
        #     if not otp_verified and  self.request.POST.get('value') and not request.POST.get('resend-otp'):
        #         print(self.request.session['vault2wallet-otp'])
        #         print(self.request.POST.get('value'))
        #         if self.request.POST.get('value') == self.request.session['vault2wallet-otp']:
        #             print('sms otp verified')
        #             request.session['vault2wallet-otp-verified'] = True
        #             sms_redirect = True
        #         else:
        #             print('sms otp invalid')
        #             return HttpResponse(json.dumps({"status":"otp-not-verified"}), content_type='application/json')

        #     elif not otp_verified:
        #         print('sms otp not verified')
        #         pin =  _get_pin(self)
        #         self.request.session['vault2wallet-otp'] = pin
        #         send_otp(self, pin, self.request.user.phone_number)
        #         return HttpResponse(json.dumps({"status":"get-otp"}), content_type='application/json')

        # if request.user.google_2fa:
        #     print('totp')
        #     if self.request.POST.get('value') and not sms_redirect:
        #         totp_code = self.request.POST.get('value', None)
        #         totp = pyotp.TOTP(self.request.user.google_2fa_key)

        #         if not totp.verify(totp_code):
        #             print('totp not verified')
        #             return HttpResponse(json.dumps({"status":"totp-not-verified"}), content_type='application/json')
        #     else:
        #         print('totp get')
        #         return HttpResponse(json.dumps({"status":"get-totp"}), content_type='application/json')
        #     print('totp verified')

        try:
            del request.session['vault2wallet-otp-verified']
        except:
            pass

        vault_address = request.session['address']

        access = create_connection()

        pre_balance = get_account_balance()

        try:
            result = access.importprivkey(vault_address,'', False)
        except:
            return HttpResponse(json.dumps({"error": "Invalid private key<br><small>The private key should be encoded in base58check using wallet import format (WIF)<small>"}), content_type='application/json')

        new_balance = balance = get_account_balance()
        amount_added = new_balance - pre_balance

        # complete_pending_transactions()

        public_key = get_publict_key(vault_address)

        VaultTransaction.objects.create(user=request.user, currency="btc", balance='', 
                        amount= amount_added, transaction_id='', transaction_to='',
                        transaction_type='from_vault', transaction_from=public_key)

        btc_limit_obj, create = MinimunCoin.objects.get_or_create(site=Site.objects.get_current())
        btc_limit_obj.low_limit_alert = True
        btc_limit_obj.save()

        return HttpResponse(json.dumps({"success": True}), content_type='application/json')


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class OrderView(View):

    def post(self, request, *args, **kwargs):

        if request.user.kyc_verified == 'not-verified':
            return JsonResponse({'status': False, 'error': _('Account not activated, Please verify kyc details')})

        if request.user.kyc_verified == 'processing':
            return JsonResponse({'status': False, 'error': _('Account not activated, Please wait for kyc approval')})

        if request.user.kyc_verified == 'rejected':
            return JsonResponse({'status': False, 'error': _('Kyc details rejected, Please contact administrator for details')})
        
        if request.POST['is_otc']=='1':
            bank_account = BankAccount.objects.filter(user=request.user)
            if not bank_account:
                return JsonResponse({'status': False, 'error': _('You need to set Bank Account to start trade')})
        form = ExchangeForm(request.POST)

        if form.is_valid():

            try:
                # sheduling mutiple request access (but only works for single server process)
                lock = Lock()
                lock.acquire()

                # processing and saving order
                form_object = form.save(commit=False)
                status = complete_order(request, form_object)

                if status == False:
                    return JsonResponse({'status': False, 'error': _('You don\'t have sufficient balance in your account to pay trading fees.')})
                
            finally:
                lock.release()

            form_object.save()
                        
            buy_order = OrderBook.objects.filter(order_type='0', trade_status=False).order_by('-price')[:20]
            sell_order = OrderBook.objects.filter(order_type='1', trade_status=False).order_by('price')[:20]

            if request.POST.get('order_type') == '0':
                if form_object.trade_status == True:
                    message = _('Buy order completed Successfully')
                else:
                    message = _('Buy order placed')
                    
                data = serializers.serialize('json', buy_order)
            else:
                if form_object.trade_status == True:
                    message = _('Sell order completed Successfully')
                else:
                    message = _('Sell order placed')

                data = serializers.serialize('json', sell_order)
            return JsonResponse({'status':True,
                                 'data':data,
                                 'message':message
                                 }, safe=False)
        else:
            errors = form.errors
            if request.POST.get('order_type') == '0':
                return JsonResponse({'status':False, 'error': errors['__all__'][0]})
            else:
                return JsonResponse({'status':False, 'error': errors['__all__'][0]})


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class AjaxDepthChartView(View):

    def get(self, request, *args, **kwargs):
        buy_order = OrderBook.objects.filter(order_type='0', trade_status=False).order_by('-price')
        sell_order = OrderBook.objects.filter(order_type='1',  trade_status=False).order_by('price')
        return HttpResponse(json.dumps({'bids': self.list_data(buy_order), 'asks': self.list_data(sell_order)}))

    def list_data(self, object):
        list_formatted_data = [[data_obj.price, data_obj.amount] for data_obj in object]
        return list_formatted_data


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class AjaxCalculateBalance(View):

    def get(self, request, *args, **kwargs):
        order_type = request.GET.get('type')
        price = request.GET.get('type')
        amount = request.GET.get('type')


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class BuyOrderList(ListView):
    """
        listing all buy orders placed by current user
    """
    context_object_name = 'buy_order_list'
    template_name = 'bitcoin/buy_order_list.html'

    def get_queryset(self, *args, **kwargs):

        buy_orders = OrderBook.objects.filter(order_type=0,user=self.request.user).order_by('-order_time')
        return buy_orders


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SellOrderList(ListView):
    """
        listing all sell orders placed by current user
    """
    context_object_name = 'sell_order_list'
    template_name = 'bitcoin/sell_order_list.html'

    def get_queryset(self, *args, **kwargs):

        sell_orders = OrderBook.objects.filter(order_type=1,user=self.request.user).order_by('-order_time')
        return sell_orders


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class DeleteOrderView(View):
    """
        delete already placed buy or sell order
    """
    def get(self, request, *args, **kwargs):

        order_object = get_object_or_404(OrderBook, pk=kwargs['pk'], user=request.user, trade_status= False)
        sgd_wallet_obj, create = SGDWallet.objects.get_or_create(user=request.user)

        if order_object.order_type == '0':
            new_balance = (order_object.amount - order_object.coins_covered) * order_object.price  + sgd_wallet_obj.amount
        
        if order_object:
            if order_object.coins_covered == 0.0:
                delted = order_object.delete()
            else:
                OrderBook.objects.filter(id=order_object.pk).update(trade_status=True, canceled=True)
            
            if order_object.order_type == '0':
                sgd_wallet_obj.amount = new_balance
            sgd_wallet_obj.save()

            return JsonResponse({'success': True})

        return JsonResponse({'success': False})

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class AcceptOrderView(TemplateView):
    template_name = 'bitcoin/accept_order.html'

    # def get(self, request, *args, **kwargs):
    #     data = self.request.GET
    #     pk = int(data['id'])
    #     tradingOrder = OrderBook.objects.filter(id=pk).select_related('user')
    #     context = {
    #         'tradingorder': tradingOrder
    #     }
    #     return render(request, 'bitcoin/accept_order.html', context)
    def get_context_data(self, **kwargs):
        context = super(AcceptOrderView, self).get_context_data(**kwargs)
        context['no_record_check'] = int(1)
        return context
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        if request.user.kyc_verified == 'not-verified':
            alert = _('Account not activated, Please verify kyc details')
            # return JsonResponse({'status': False, 'error': _('Account not activated, Please verify kyc details')})
        if request.user.kyc_verified == 'processing':
            alert = _('Account not activated, Please wait for kyc approval')
            # return JsonResponse({'status': False, 'error': _('Account not activated, Please wait for kyc approval')})
        if request.user.kyc_verified == 'rejected':
            alert = _('Kyc details rejected, Please contact administrator for details')
            # return JsonResponse({'status': False, 'error': _('Kyc details rejected, Please contact administrator for details')})
        bank_account = BankAccount.objects.filter(user=request.user)
        if not bank_account:
            alert = _('You need to set Bank Account to start trade')
            # return JsonResponse({'status': False, 'error': _('You need to set Bank Account to start trade')})
            # return render(request, 'settings.html', {'status': False, 'error': _('You need to set Bank Account to start trade')})
        pk = request.POST['id'] # this is order id
        to_amount = request.POST['amount'] # this is pay amount
        tradingOrder = OrderBook.objects.filter(id=pk).select_related('user')
        from_amount = tradingOrder.all().first().price*float(to_amount)
        request_user = request.user.id
        order_type = tradingOrder.all().first().order_type
        if order_type == '1':
            user_id = tradingOrder.all().first().user_id
            bank_account = BankAccount.objects.filter(user=user_id)
            paypal_account = PaypalAccount.objects.filter(user=user_id)
            wechat_account = WechatAccount.objects.filter(user=user_id)
            alipay_account = AlipayAccount.objects.filter(user=user_id)
            phone_number = User.objects.filter(id=user_id).values_list('phone_number').first()
            
            pending_obj = PendingOrder(user_id=request_user, order_id = pk)
            pending_obj.pending_amount = float(to_amount)
            pending_obj.save()
            print (pending_obj.id)
            try:
                context = {
                    'tradingorder': tradingOrder,
                    'to_amount': to_amount,
                    'from_amount': from_amount,
                    'bank_accounts': bank_account,
                    'paypal_accounts': paypal_account,
                    'wechat_accounts': wechat_account,
                    'alipay_accounts': alipay_account,
                    'request_user': request_user,
                    'phone_number': phone_number,
                    'alert': alert,
                    'pending_obj': pending_obj
                }
            except :
                context = {
                    'tradingorder': tradingOrder,
                    'to_amount': to_amount,
                    'from_amount': from_amount,
                    'bank_accounts': bank_account,
                    'paypal_accounts': paypal_account,
                    'wechat_accounts': wechat_account,
                    'alipay_accounts': alipay_account,
                    'phone_number': phone_number,
                    'request_user': request_user,
                    'pending_obj': pending_obj.id
                }
        else :
            sender = request.POST['sender']
            sendername = User.objects.filter(id=sender).first().username
            bank_account = BankAccount.objects.filter(user=sender)
            paypal_account = PaypalAccount.objects.filter(user=sender)
            wechat_account = WechatAccount.objects.filter(user=sender)
            alipay_account = AlipayAccount.objects.filter(user=sender)
            context = {
                'tradingorder': tradingOrder,
                'to_amount': to_amount,
                'from_amount': from_amount,
                'bank_accounts': bank_account,
                'paypal_accounts': paypal_account,
                'wechat_accounts': wechat_account,
                'alipay_accounts': alipay_account,
                'request_user': request_user,
                'email': sendername,
                'sender': sender
            }
            
        return render(request, 'bitcoin/accept_order.html', context)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ConfirmFiatTransactionsView(View):
    template_name = 'bitcoin/confirm_fiat_view.html'

    def get(self, request, *args, **kwargs):
        fiat_txs = ConfirmFiatTransaction.objects.filter(is_confirm=0, is_cancel=0, receiver=request.user.id).select_related('sender')
        context = {
            'fiat_txs': fiat_txs
        }
        return render(request, self.template_name, context)
    def post(self, request, *args, **kwargs):
        form = ConfirmFiatTransactionForm(request.POST)
        form = form.save(commit=False)
        form.save()
        return JsonResponse({'success': True})

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class ConfirmRequestFiatTransactionsView(View):
    template_name = 'bitcoin/confirm_request_view.html'

    def get(self, request, *args, **kwargs):
        request_crypto_txs = ConfirmCrytpoRequest.objects.filter(is_confirm=0, is_cancel=0, receiver=request.user.id).select_related('sender')
        context = {
            'request_crypto_txs': request_crypto_txs
        }
        return render(request, self.template_name, context)
    def post(self, request, *args, **kwargs):
        if request.POST['is_otc']=='1':
            form = ConfirmCrytpoRequestForm(request.POST)
            if form.is_valid():
                form.save()
                return JsonResponse({'status':True, 'error':form.errors})
            else :
                return JsonResponse({'status':False, 'error':form.errors})
        else :
            form = OrderCoverTransactionForm(request.POST)
            if form.is_valid():
                form.save()
                OrderBook.objects.filter(pk=request.POST['order']).update(coins_covered=F('coins_covered')+request.POST['amount'])
                return JsonResponse({'status':True, 'error':form.errors})
            else :
                return JsonResponse({'status':False, 'error':form.errors})

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class CancelRequestFiatTransactionsView(View):
    def post(self, request, *args, **kwargs):
        ConfirmCrytpoRequest.objects.filter(pk=request.POST['id']).update(is_cancel=True)
        return JsonResponse({'success': True})

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class GetUserBalanceTransactionView(UserPassesTestMixin, View):
    def test_func(self):
        if not self.request.user.is_superuser:
            raise PermissionDenied()
        return True

    def post(self, request, *args, **kwargs):
        user = User.objects.filter(email=request.POST['useremail']).first()
        # access = AuthServiceProxy("http://tixoner:123456789tixon@68.183.176.17:9999")
        # transactions = access.listtransactions("*") #self.request.user.username 
        # context['btctransactions'] = [txn for txn in transactions if txn['category'] == 'receive']
        # context['btcsend_list'] = Transaction.objects.filter(user=self.request.user, currency="BTC", transaction_type='withdrawal')

        access = AuthServiceProxy("http://tixoner:123456789tixon@68.183.185.230:9988")
        txch_transactions = access.listtransactions(request.POST['useremail'])
        txch_transactions = [txn for txn in txch_transactions if txn['category'] == 'move']
        try:
            data = {
                'btc_balance': get_balance(user, 'BTC'),
                'txch_balance': get_balance(user, 'TXCH'),
                'txch_transactions': txch_transactions
            }
            return JsonResponse({'success': True, 'user_balance': data})    
        except:
            return JsonResponse({'success': False})

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class CancelAcceptFiatTransactionsView(View):
    def post(self, request, *args, **kwargs):
        ConfirmFiatTransaction.objects.filter(pk=request.POST['id']).update(is_cancel=True)
        return JsonResponse({'success': True})

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class UpdateOtcOrderView(View):
    def post(self, request, *args, **kwargs):
        ConfirmFiatTransaction.objects.filter(pk=request.POST['id']).update(is_confirm=True)
        order_id = ConfirmFiatTransaction.objects.filter(pk=request.POST['id']).all().first().order_id
        crypto_amount = ConfirmFiatTransaction.objects.filter(pk=request.POST['id']).all().first().crypto_amount
        crypto_amount += OrderBook.objects.filter(pk=order_id).all().first().coins_covered
        OrderBook.objects.filter(pk=order_id).update(coins_covered=crypto_amount)
        return JsonResponse({'success': True})


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class NotificationListView(View):
    """
        Listing all notifications of current user
    """
    def get(self, request, *args, **kwargs):
        
        notification_count = NotificationUser.objects.filter(user=self.request.user, is_readed=False).count()

        NotificationUser.objects.filter(user=self.request.user).update(is_readed=True)

        context = {
            'notifications': Notification.objects.filter(user= self.request.user).order_by('-date'),
            'notification_count': notification_count
        }

        return render(request, 'bitcoin/notifications.html', context) 

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class DownloadReport(View):
    """
        retrive deposit and withdrawal of a user in xlsx format
    """
    def get(self, request, *args, **kwargs):
        # rendering ReportForm in template
        context ={
            'form': ReportForm()
        }
        return render(request, 'bitcoin/report.html', context)

    def post(self, request, *args, **kwargs):
        # filtering Transaction object using date rage and generating xlsx sheet
        form = ReportForm(request.POST)

        #form validation
        if form.is_valid():
            start_date = form.cleaned_data['fromdate']
            end_date = form.cleaned_data['todate']+ timedelta(days=1)
        else:
            context ={
                'form': form
            }
            return render(request, 'bitcoin/report.html', context)

        # queryset transactions in date rage
        transactions = Transaction.objects.filter(user=request.user, pending=False, transaction_type__in=['withdrawal','deposit'], 
            date__range=[start_date,end_date]).order_by('date')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="somefilename.xlsx"'

        wb = Workbook()
        ws = wb.active

        row_num = 0

        columns = [
            (u"Date", 15),
            (u"Type", 70),
            (u"Currency", 70),
            (u"Amount", 70),
            (u"Transaction Fee",70)
        ]
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            c.font = Font(bold=True)

        for obj in transactions:
            row_num += 1
            row = [
                obj.date.strftime('%d/%m/%y %H:%M:%S'),
                obj.transaction_type.capitalize(),
                obj.currency.upper(),
                obj.amount,
                obj.transaction_fee
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.alignment = Alignment(wrapText=True)

        wb.save(response)

        return response


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class SGDWalletsView(ListView):
    """
        list all user sgd wallets
    """
    context_object_name = 'sgd_wallet_list'
    template_name = 'bitcoin/sgd-wallet-list.html'

    def get_queryset(self, *args, **kwargs):

        sgd_wallet_list = SGDWallet.objects.all()

        q = self.request.GET.get('q', None)

        if q:
            sgd_wallet_list = sgd_wallet_list.filter(Q(user__username__icontains=q))

        return sgd_wallet_list


@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class AddSGDView(View):
    """
        updating user wallet amount
    """
    template_name ='bitcoin/add_sgd.html'
    form_class = AddSGDForm

    def get(self, request, *args, **kwargs):
        wallet = SGDWallet.objects.get(pk=kwargs['pk'])
        form = AddSGDForm()

        context ={
            'wallet': wallet,
            'form': form
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        wallet = SGDWallet.objects.get(pk=kwargs['pk'])
        form = AddSGDForm(request.POST)

        if form.is_valid():
            amount = float(form.cleaned_data['amount'])
            wallet.amount += amount
            wallet.save()

            messages.success(request, 'SGD Wallet updated of user %s' %wallet.user.username)
            return redirect(reverse('coins:sgd_wallets'))
        
        context = {
            'form': form,
            'wallet': wallet
        }
        return render(request, self.template_name, context)

class AjaxOrderBookView(View):
    def get(self, request, *args, **kwargs):
        buy_order = OrderBook.objects.filter(Q(order_type='0'), Q(trade_status=False), Q(amount__gt = F('coins_covered'))).order_by('-price')[:20]
        sell_order = OrderBook.objects.filter(Q(order_type='1'), Q(trade_status=False), Q(amount__gt = F('coins_covered'))).order_by('price')[:20]
        pending_order = PendingOrder.objects.filter(Q(is_pending=True)).values('order_id').annotate(sum_pending_amount=Sum('pending_amount'))
        json_buy_order = serializers.serialize('json', buy_order)
        json_sell_order = serializers.serialize('json', sell_order)
        # json_pending_order = serializers.serialize('json', list(pending_order), fields=('order_id',))
        return JsonResponse({
            'json_buy_order': json_buy_order,
            'json_sell_order': json_sell_order,
            'json_pending_order': json.dumps(list(pending_order))
        }, safe=False)

@method_decorator(login_required, name='dispatch')
@method_decorator(check_otp, name='dispatch')
@method_decorator(check_2fa, name='dispatch')
class CancelPendingOrderView(View):
    def get(self, request, *args, **kwargs):
        PendingOrder.objects.filter(id=self.kwargs['pk']).update(is_pending=False) 
        return redirect(reverse('welcomeotc'))